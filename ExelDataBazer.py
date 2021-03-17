from openpyxl import load_workbook


book = load_workbook('DataBaze.xlsx')
sheet_1 = book['Тишина']
sticker_page = book['Тишина']

print(book.worksheets)
for i in range(1, 4):
    print(sticker_page.cell(row=1, column=i).value)
