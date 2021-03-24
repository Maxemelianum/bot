from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook
import random



TOKEN = '1591047164:AAGvlzpApFB-u8qpGZYE1YQmKIjL9lotSLs'
book = load_workbook('DataBaze.xlsx')
sheet_1 = book['Тишина']
sticker_page = book['Тишина']


what = [
    'Идиот, я же не понимаю нихрена!',
    'А команды тебе просто так даны?',
    'У тебя IQ как у хлебушка! Используй команды!'

]

hi = [
    'Ну привет кожанный.',
    'Готовь свои нервы.',
    'Удачи, кожанный ублюдок.'
]

helper = [
    '/start - запуск бота\n/user - посмотреть инфо о себе\n/stop - остановить бота\nНо тебе уже ничего не поможет(',
    'Ты жалок! Ты просишь помщи у бота!\n/start - запуск бота\n/user - посмотреть инфо о себе\n/stop - остановить бота',
    'Сам разбирайся кожанный )'
]
stoping = [
    'Ты уверен?',
    'Иди поспи.',
    'Кожанный мешок хочет остановить бота. Хорошая попытка.'

]
stick = [
    'Вот твое фото',
    'Фоточка тобi на',
    'живи с этим'
]

randomizer = random.SystemRandom()


def main():
    updater = Updater(token=TOKEN)  # объект, который ловит сообщения от Telegrama

    dispather = updater.dispatcher

    handler_help = CommandHandler('help', do_bulling)
    start_handler = CommandHandler('start', do_start)
    stop_handler = CommandHandler('stop', do_stop)
    name = CommandHandler('user', do_user)
    sticker = CommandHandler('img', do_text)
    attack = CommandHandler('attack', do_attack)
    some = MessageHandler(Filters.text, do_some)
    stick_handler = MessageHandler(Filters.sticker, do_sticker)
    handlerbull = MessageHandler(Filters.all, do_echo)

    dispather.add_handler(start_handler)
    dispather.add_handler(handler_help)
    dispather.add_handler(name)
    dispather.add_handler(attack)
    dispather.add_handler(stick_handler)
    dispather.add_handler(stop_handler)
    dispather.add_handler(sticker)
    dispather.add_handler(some)
    dispather.add_handler(handlerbull)

    updater.start_polling()
    updater.idle()


def do_echo(update, context):
    update.message.reply_text(randomizer.choice(what))


def do_start(update, context):

    update.message.reply_text(randomizer.choice(hi))


def do_bulling(update, context):
    keyboard = [['/help', '/user', '/stop'],
                ['/img']]
    update.message.reply_text(randomizer.choice(helper),
    reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))


def do_stop(update, context):
    update.message.reply_text(randomizer.choice(stoping))


def do_user(update: Update, context):
    user_id = update.message.from_user.id
    name = update.message.from_user.name
    update.message.reply_text(text=f'Твой disgusting name:{name}\nТвой disgusting id:{user_id}')


def do_text(update, context):
        img = 'https://imbt.ga/dBmlYLNaCQ'
        update.message.reply_text(randomizer.choice(stick))
        update.message.reply_text(f'{img}')


def do_some(update: Update, context):
    text = update.message.text

    for row in range(2, sticker_page.max_row + 1):
        catch_fase = sticker_page.cell(row=row, column=4).value
        print(text)
        print(catch_fase)
        if catch_fase in text:
            sticker_id = sticker_page.cell(row=row, column=3).value
            update.message.reply_sticker(sticker_id)

    if text == 'Нет':
        update.message.reply_text('****** ответ', reply_markup=ReplyKeyboardRemove())
        update.message.reply_sticker('CAACAgIAAxkBAAECEd1gUd5kmaGDSLkdJEidOZwxCp7C_AACGAADoF_dLYLQt9fbwjB_HgQ')
    elif text == 'Да':
        update.message.reply_text('Вы использовали команду /user', reply_markup=ReplyKeyboardRemove())
    elif text == 'Давай не будем':
        update.message.reply_text('Вы использовали команду /stop', reply_markup=ReplyKeyboardRemove())
    elif text == 'Хватит':
        update.message.reply_text('Вы использовали команду /img', reply_markup=ReplyKeyboardRemove())
    else:
        update.message.reply_text(randomizer.choice(what), reply_markup=ReplyKeyboardRemove())


def do_sticker(update: Update, context):
    sticker_id = update.message.sticker.file_id
    update.message.reply_sticker(sticker_id)


def do_attack(update, context):
    keyboard = [['Да'], ['Нет']]
    img = 'https://imbt.ga/TlRYeptk2h'
    update.message.reply_text(f'Начать атаку?{img}',
    reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))


main()